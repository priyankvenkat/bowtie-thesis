# Node and edge metrics + export
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from sentence_transformers import SentenceTransformer, util
from .graph_utils import normalize
from .constants import CSV_LOG
import os 
import re 
from openpyxl import load_workbook, Workbook
import torch
import numpy as np
from pathlib import Path

model = SentenceTransformer('all-MiniLM-L6-v2')

def map_pred_to_gt_nodes(pred_nodes, gt_nodes, threshold=0.8):
    from sentence_transformers import SentenceTransformer, util
    from .graph_utils import normalize
    import torch

    model = SentenceTransformer("all-MiniLM-L6-v2")  # or reuse your loaded one

    norm_pred = [normalize(p) for p in pred_nodes]
    norm_gt = [normalize(g) for g in gt_nodes]

    emb_pred = model.encode(norm_pred, convert_to_tensor=True)
    emb_gt = model.encode(norm_gt, convert_to_tensor=True)

    sim_matrix = util.pytorch_cos_sim(emb_pred, emb_gt).cpu().numpy()

    mapping = {}
    matched_gt = set()

    for i, pred_node in enumerate(norm_pred):
        j = sim_matrix[i].argmax()
        if sim_matrix[i][j] >= threshold and norm_gt[j] not in matched_gt:
            mapping[pred_node] = norm_gt[j]
            matched_gt.add(norm_gt[j])
        else:
            mapping[pred_node] = pred_node  # no remap fallback

    return mapping


def evaluate_with_soft_remap(gt_edges, pred_edges, gt_roles, pred_roles, threshold=0.85):
    """
    Performs edge metric evaluation after remapping predicted nodes to ground truth nodes
    using cosine similarity (based on node labels), without altering manually assigned roles.

    Returns:
        edge_metrics_rows: list of dicts (Edge Type breakdown)
        remapped_edges: list of (src, tgt) tuples
        remapped_roles: dict of remapped predicted roles
        node_mapping: dict of pred_node -> matched_gt_node
    """


    # Extract nodes
    gt_nodes = {normalize(n) for e in gt_edges for n in e}
    pred_nodes = {normalize(n) for e in pred_edges for n in e}

    # Step 1: Generate mapping
    node_mapping = map_pred_to_gt_nodes(pred_nodes, gt_nodes, threshold=threshold)

    # Step 2: Remap predicted edges
    remapped_edges = [
        (node_mapping.get(normalize(a), a), node_mapping.get(normalize(b), b))
        for a, b in pred_edges
    ]

    # Step 3: Remap predicted roles
    remapped_roles = {
        node_mapping.get(normalize(k), k): v for k, v in pred_roles.items()
    }

    # Step 4: Run edge metrics
    edge_metrics_rows = compute_edge_metrics(gt_edges, remapped_edges, gt_roles, remapped_roles)

    return edge_metrics_rows, remapped_edges, remapped_roles, node_mapping


def compute_node_metrics(gt_edges, pred_edges, gt_roles, pred_roles, threshold=0.85):
    from collections import defaultdict

    role_categories = ["Cause", "Mechanism", "Central Event", "Barrier", "Consequence"]
    role_metrics = {}

    # Normalize role keys once
    gt_roles = {normalize(k): v for k, v in gt_roles.items()}
    pred_roles = {normalize(k): v for k, v in pred_roles.items()}

    # Group nodes by role
    gt_by_role = {role: set() for role in role_categories}
    pred_by_role = {role: set() for role in role_categories}

    for e in gt_edges:
        for node in e:
            node = normalize(node)
            role = gt_roles.get(node)
            if role in gt_by_role:
                gt_by_role[role].add(node)

    for e in pred_edges:
        for node in e:
            node = normalize(node)
            role = pred_roles.get(node)
            if role in pred_by_role:
                pred_by_role[role].add(node)

    # Optional: detect nodes with multiple predicted roles
    node_to_roles = defaultdict(set)
    for role, nodes in pred_by_role.items():
        for node in nodes:
            node_to_roles[node].add(role)
    multi_role_nodes = {n for n, r in node_to_roles.items() if len(r) > 1}
    if multi_role_nodes:
        print("⚠️ Warning: Multi-role predicted nodes detected:", multi_role_nodes)

    # Matching logic
    total_tp = total_fp = total_fn = 0
    total_role_mismatch = 0

    for role in role_categories:
        gt_nodes = list(gt_by_role[role])
        pred_nodes = list(pred_by_role[role])

        if not gt_nodes and not pred_nodes:
            continue

        if not gt_nodes or not pred_nodes:
            tp, fp, fn = 0, len(pred_nodes), len(gt_nodes)
        else:
            emb_gt = model.encode(gt_nodes, convert_to_tensor=True)
            emb_pred = model.encode(pred_nodes, convert_to_tensor=True)
            sim_matrix = util.pytorch_cos_sim(emb_pred, emb_gt).cpu().numpy()

            matched_gt = set()
            matched_pred = set()
            tp = role_mismatches = 0

            for i, pred_node in enumerate(pred_nodes):
                j = sim_matrix[i].argmax()
                max_sim = sim_matrix[i][j]
                gt_node = gt_nodes[j]
                if max_sim >= threshold and gt_node not in matched_gt and pred_node not in matched_pred:
                    gt_role = gt_roles.get(gt_node)
                    pred_role = pred_roles.get(pred_node)
                    if gt_role == pred_role:
                        tp += 1
                    else:
                        role_mismatches += 1
                    matched_gt.add(gt_node)
                    matched_pred.add(pred_node)

            fp = len(pred_nodes) - tp - role_mismatches
            fn = len(gt_nodes) - tp - role_mismatches
            total_role_mismatch += role_mismatches

        precision = tp / (tp + fp) if tp + fp > 0 else 0.0
        recall = tp / (tp + fn) if tp + fn > 0 else 0.0
        f1 = 2 * precision * recall / (precision + recall) if precision + recall > 0 else 0.0
        jaccard = tp / (tp + fp + fn) if tp + fp + fn > 0 else 0.0

        role_metrics[role] = {
            "Precision": round(precision, 3),
            "Recall": round(recall, 3),
            "F1": round(f1, 3),
            "Jaccard": round(jaccard, 3),
            "TP": tp, "FP": fp, "FN": fn,
            "Role Mismatches": role_mismatches
        }

        total_tp += tp
        total_fp += fp
        total_fn += fn

    # Overall metrics
    precision = total_tp / (total_tp + total_fp) if total_tp + total_fp > 0 else 0.0
    recall = total_tp / (total_tp + total_fn) if total_tp + total_fn > 0 else 0.0
    f1 = 2 * precision * recall / (precision + recall) if precision + recall > 0 else 0.0
    jaccard = total_tp / (total_tp + total_fp + total_fn) if total_tp + total_fp + total_fn > 0 else 0.0

    role_metrics["Overall"] = {
        "Precision": round(precision, 3),
        "Recall": round(recall, 3),
        "F1": round(f1, 3),
        "Jaccard": round(jaccard, 3),
        "TP": total_tp, "FP": total_fp, "FN": total_fn,
        "Role Mismatches": total_role_mismatch
    }

    return role_metrics

def export_node_metrics_to_excel(metrics_dict, method, prompt, domain, model, central_event, file_path=CSV_LOG):
    import os
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Convert dict to DataFrame
    df = pd.DataFrame.from_dict(metrics_dict, orient="index").reset_index().rename(columns={"index": "Role"})

    # Add metadata columns
    df["Method"] = method
    df["Prompt"] = prompt
    df["Domain"] = domain
    df["Model"] = model
    df["Central Event"] = central_event

    # Load or create workbook
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Get or create worksheet
    sheet_name = "Node_Metrics"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_empty = ws.max_row <= 1
    else:
        ws = wb.create_sheet(sheet_name)
        is_empty = True

    # Add a blank row to separate groups
    if not is_empty:
        ws.append([])

    # Write rows
    for r in dataframe_to_rows(df, index=False, header=is_empty):
        ws.append(r)

    wb.save(file_path)
    return df


def update_csv(row, file_path=CSV_LOG):
    import pandas as pd
    from openpyxl import Workbook
    import openpyxl

    ged_df = pd.DataFrame([row])
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "GED_Metrics" in wb.sheetnames:
        ws1 = wb["GED_Metrics"]
        is_empty = ws1.max_row <= 1
    else:
        ws1 = wb.create_sheet("GED_Metrics")
        is_empty = True

    if not is_empty:
        ws1.append([])

    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(ged_df, index=False, header=is_empty):
        ws1.append(r)

    # (Optional) Write Node_Mappings sheet here if needed

    wb.save(file_path)
    return pd.read_excel(file_path, sheet_name="GED_Metrics")

def compute_edge_metrics(gt_edges, pred_edges, gt_roles, pred_roles):
    from collections import defaultdict

    # Normalize role dictionaries
    gt_roles = {normalize(k): v for k, v in gt_roles.items()}
    pred_roles = {normalize(k): v for k, v in pred_roles.items()}

    # Normalize edges
    norm_gt = [(normalize(a), normalize(b)) for a, b in gt_edges]
    norm_pred = [(normalize(a), normalize(b)) for a, b in pred_edges]

    # Detect nodes with multiple predicted roles
    node_to_roles = defaultdict(set)
    for node, role in pred_roles.items():
        node_to_roles[node].add(role)
    multi_role_nodes = {n for n, r in node_to_roles.items() if len(r) > 1}
    if multi_role_nodes:
        print("⚠️ Warning: Multi-role predicted nodes detected:", multi_role_nodes)

    # Helper to determine edge type safely
    def edge_type(a, b):
        role_a = gt_roles.get(a) or pred_roles.get(a) or "Unknown"
        role_b = gt_roles.get(b) or pred_roles.get(b) or "Unknown"
        if a in multi_role_nodes or b in multi_role_nodes:
            return "Ambiguous"
        return f"{role_a}->{role_b}"

    # Classify edges
    gt_classified = {(a, b): edge_type(a, b) for (a, b) in norm_gt}
    pred_classified = {(a, b): edge_type(a, b) for (a, b) in norm_pred}

    metrics = {}
    role_mismatches = 0
    overall_tp = overall_fp = overall_fn = 0

    # Count TP and FP
    for edge, pred_type in pred_classified.items():
        gt_type = gt_classified.get(edge)
        if edge in gt_classified:
            if gt_type == pred_type:
                # Perfect match
                metrics.setdefault(pred_type, {"TP": 0, "FP": 0, "FN": 0, "Role Mismatch": 0})
                metrics[pred_type]["TP"] += 1
                overall_tp += 1
            else:
                # Role mismatch (same nodes, wrong type)
                metrics.setdefault(pred_type, {"TP": 0, "FP": 0, "FN": 0, "Role Mismatch": 0})
                metrics[pred_type]["Role Mismatch"] += 1
                role_mismatches += 1
                overall_fp += 1
        else:
            # Extra predicted edge
            metrics.setdefault(pred_type, {"TP": 0, "FP": 0, "FN": 0, "Role Mismatch": 0})
            metrics[pred_type]["FP"] += 1
            overall_fp += 1

    # Count FN
    for edge, gt_type in gt_classified.items():
        if edge not in pred_classified or pred_classified.get(edge) != gt_type:
            metrics.setdefault(gt_type, {"TP": 0, "FP": 0, "FN": 0, "Role Mismatch": 0})
            metrics[gt_type]["FN"] += 1
            overall_fn += 1

    # Convert to rows
    rows = []
    for etype, vals in metrics.items():
        tp, fp, fn, mismatch = vals["TP"], vals["FP"], vals["FN"], vals["Role Mismatch"]
        prec = tp / (tp + fp) if tp + fp > 0 else 0.0
        rec = tp / (tp + fn) if tp + fn > 0 else 0.0
        f1 = 2 * prec * rec / (prec + rec) if prec + rec > 0 else 0.0
        jacc = tp / (tp + fp + fn) if tp + fp + fn > 0 else 0.0
        rows.append({
            "Edge Type": etype,
            "Precision": round(prec, 3),
            "Recall": round(rec, 3),
            "F1": round(f1, 3),
            "Jaccard": round(jacc, 3),
            "TP": tp, "FP": fp, "FN": fn,
            "Role Mismatches": mismatch
        })

    # Overall
    prec = overall_tp / (overall_tp + overall_fp) if overall_tp + overall_fp > 0 else 0.0
    rec = overall_tp / (overall_tp + overall_fn) if overall_tp + overall_fn > 0 else 0.0
    f1 = 2 * prec * rec / (prec + rec) if prec + rec > 0 else 0.0
    jacc = overall_tp / (overall_tp + overall_fp + overall_fn) if overall_tp + overall_fp + overall_fn > 0 else 0.0

    rows.append({
        "Edge Type": "Overall",
        "Precision": round(prec, 3),
        "Recall": round(rec, 3),
        "F1": round(f1, 3),
        "Jaccard": round(jacc, 3),
        "TP": overall_tp, "FP": overall_fp, "FN": overall_fn,
        "Role Mismatches": role_mismatches
    })

    return rows

def export_edge_metrics_to_excel(rows, method, prompt, domain, model, central_event, file_path=CSV_LOG):
    import os
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Convert row list to DataFrame
    df = pd.DataFrame(rows)

    # Add metadata
    df["Method"] = method
    df["Prompt"] = prompt
    df["Domain"] = domain
    df["Model"] = model
    df["Central Event"] = central_event

    # Load or create workbook
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Get or create worksheet
    sheet_name = "Edge_Metrics"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_empty = ws.max_row <= 1
    else:
        ws = wb.create_sheet(sheet_name)
        is_empty = True

    # Append a blank row to visually separate metric blocks
    if not is_empty:
        ws.append([])

    # Append rows
    for r in dataframe_to_rows(df, index=False, header=is_empty):
        ws.append(r)

    wb.save(file_path)
    return df

def export_detailed_edge_comparisons(gt_edges, pred_edges, gt_roles, pred_roles, method, prompt, domain, model, central_event, file_path="ged_results.xlsx"):
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import os
    import pandas as pd

    gt_roles = {normalize(k): v for k, v in gt_roles.items()}
    pred_roles = {normalize(k): v for k, v in pred_roles.items()}

    gt_set = set((normalize(a), normalize(b)) for a, b in gt_edges)
    pred_set = set((normalize(a), normalize(b)) for a, b in pred_edges)

    detailed_rows = []

    def get_role(a, b):
        return gt_roles.get(a) or pred_roles.get(a, "Unknown"), gt_roles.get(b) or pred_roles.get(b, "Unknown")

    # TP
    for a, b in gt_set & pred_set:
        role_a, role_b = get_role(a, b)
        detailed_rows.append({
            "Source": a, "Target": b,
            "Edge Type": f"{role_a}->{role_b}",
            "Match Status": "TP"
        })

    # FN
    for a, b in gt_set - pred_set:
        role_a, role_b = get_role(a, b)
        detailed_rows.append({
            "Source": a, "Target": b,
            "Edge Type": f"{role_a}->{role_b}",
            "Match Status": "FN"
        })

    # FP
    for a, b in pred_set - gt_set:
        role_a = pred_roles.get(a) or gt_roles.get(a, "Unknown")
        role_b = pred_roles.get(b) or gt_roles.get(b, "Unknown")
        detailed_rows.append({
            "Source": a, "Target": b,
            "Edge Type": f"{role_a}->{role_b}",
            "Match Status": "FP"
        })

    for row in detailed_rows:
        row.update({
            "Method": method, "Prompt": prompt,
            "Domain": domain, "Model": model,
            "Central Event": central_event
        })

    df = pd.DataFrame(detailed_rows)

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheet = "Edge_Comparisons"
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
    is_empty = ws.max_row <= 1

    if not is_empty:
        ws.append([])

    for r in dataframe_to_rows(df, index=False, header=is_empty):
        ws.append(r)

    wb.save(file_path)
    return df

def export_node_comparisons_to_excel(gt_nodes, pred_nodes, gt_roles, pred_roles,
                                      method, prompt, model_name, domain,
                                      central_event, file_path="ged_results.xlsx", threshold=0.85):
    norm_gt_nodes = [normalize(n) for n in gt_nodes]
    norm_pred_nodes = [normalize(n) for n in pred_nodes]

    emb_gt = model.encode(norm_gt_nodes, convert_to_tensor=True)
    emb_pred = model.encode(norm_pred_nodes, convert_to_tensor=True)

    sim_matrix = util.pytorch_cos_sim(emb_pred, emb_gt).cpu().numpy()
    matched_gt = set()
    rows = []

    for i, pred_label in enumerate(norm_pred_nodes):
        j = sim_matrix[i].argmax()
        max_sim = sim_matrix[i][j]
        gt_label = norm_gt_nodes[j]

        if max_sim >= threshold and gt_label not in matched_gt:
            matched_gt.add(gt_label)
            gt_role = gt_roles.get(gt_label, "")
            pred_role = pred_roles.get(pred_label, "")
            match_status = "TP" if gt_role == pred_role else "Role Mismatch"
        else:
            gt_label = ""
            gt_role = ""
            pred_role = pred_roles.get(pred_label, "")
            match_status = "FP"

        rows.append({
            "Node Label": pred_label,
            "Ground Truth Role": gt_role,
            "Predicted Role": pred_role,
            "Match Status": match_status,
            "Method": method,
            "Prompt": prompt,
            "Model": model_name,
            "Domain": domain,
            "Central Event": central_event
        })

    for i, gt_label in enumerate(norm_gt_nodes):
        if gt_label not in matched_gt:
            rows.append({
                "Node Label": gt_label,
                "Ground Truth Role": gt_roles.get(gt_label, ""),
                "Predicted Role": "",
                "Match Status": "FN",
                "Method": method,
                "Prompt": prompt,
                "Model": model_name,
                "Domain": domain,
                "Central Event": central_event
            })

    df = pd.DataFrame(rows)

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if "Node_Comparisons" in wb.sheetnames:
        ws = wb["Node_Comparisons"]
        is_empty = ws.max_row <= 1
    else:
        ws = wb.create_sheet("Node_Comparisons")
        is_empty = True

    if not is_empty:
        ws.append([])

    for r in dataframe_to_rows(df, index=False, header=is_empty):
        ws.append(r)

    wb.save(file_path)
    return df


# def export_node_comparisons_to_excel(gt_nodes, pred_nodes, gt_roles, pred_roles,
#                                       method, prompt, model_name, domain, central_event,
#                                       file_path="ged_results.xlsx", threshold=0.85):
#     from openpyxl import load_workbook, Workbook
#     from openpyxl.utils.dataframe import dataframe_to_rows
#     import os
#     import pandas as pd

#     gt_roles = {normalize(k): v for k, v in gt_roles.items()}
#     pred_roles = {normalize(k): v for k, v in pred_roles.items()}

#     norm_gt_nodes = [normalize(n) for n in gt_nodes]
#     norm_pred_nodes = [normalize(n) for n in pred_nodes]

#     emb_gt = model.encode(norm_gt_nodes, convert_to_tensor=True)
#     emb_pred = model.encode(norm_pred_nodes, convert_to_tensor=True)

#     sim_matrix = util.pytorch_cos_sim(emb_pred, emb_gt).cpu().numpy()
#     matched_gt = set()
#     rows = []

#     for i, pred_label in enumerate(norm_pred_nodes):
#         j = sim_matrix[i].argmax()
#         max_sim = sim_matrix[i][j]
#         gt_label = norm_gt_nodes[j]

#         if max_sim >= threshold and gt_label not in matched_gt:
#             matched_gt.add(gt_label)
#             gt_role = gt_roles.get(gt_label, "")
#             pred_role = pred_roles.get(pred_label, "")
#             match_status = "TP" if gt_role == pred_role else "Role Mismatch"
#         else:
#             gt_label = ""
#             gt_role = ""
#             pred_role = pred_roles.get(pred_label, "")
#             match_status = "FP"

#         rows.append({
#             "Node Label": pred_label,
#             "Ground Truth Role": gt_role,
#             "Predicted Role": pred_role,
#             "Match Status": match_status,
#             "Method": method, "Prompt": prompt,
#             "Model": model_name, "Domain": domain
#         })

#     for gt_label in norm_gt_nodes:
#         if gt_label not in matched_gt:
#             rows.append({
#                 "Node Label": gt_label,
#                 "Ground Truth Role": gt_roles.get(gt_label, ""),
#                 "Predicted Role": "",
#                 "Match Status": "FN",
#                 "Method": method, "Prompt": prompt,
#                 "Model": model_name, "Domain": domain
#                 })

#     df = pd.DataFrame(rows)

#     if os.path.exists(file_path):
#         wb = load_workbook(file_path)
#     else:
#         wb = Workbook()
#         wb.remove(wb.active)

#     sheet = "Node_Comparisons"
#     ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
#     is_empty = ws.max_row <= 1

#     if not is_empty:
#         ws.append([])

#     for r in dataframe_to_rows(df, index=False, header=is_empty):
#         ws.append(r)

#     wb.save(file_path)
#     return df


def export_cosine_mappings_to_ged_sheet(mapping_scores, method, prompt, domain, model, central_event, file_path="ged_results.xlsx"):
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import os
    import pandas as pd

    df = pd.DataFrame(mapping_scores)
    df["Method"] = method
    df["Prompt"] = prompt
    df["Domain"] = domain
    df["Model"] = model
    df["Central Event"] = central_event

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheet = "Cosine_Remapping"
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
    is_empty = ws.max_row <= 1

    if not is_empty:
        ws.append([])

    for r in dataframe_to_rows(df, index=False, header=is_empty):
        ws.append(r)

    wb.save(file_path)
    return df
