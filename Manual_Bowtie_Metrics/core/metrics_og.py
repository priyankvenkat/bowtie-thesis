# Node and edge metrics + export
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from sentence_transformers import SentenceTransformer, util
from .graph_utils import normalize
from .constants import CSV_LOG
import os 
import re 
from openpyxl import load_workbook, Workbook
import torch
import numpy as np
from pathlib import Path

model = SentenceTransformer('all-MiniLM-L6-v2')

def compute_node_metrics(gt_edges, pred_edges, gt_roles, pred_roles, threshold=0.8):
    role_categories = ["Cause", "Mechanism", "Central Event", "Barrier", "Consequence"]
    role_metrics = {}
    gt_by_role = {role: set() for role in role_categories}
    pred_by_role = {role: set() for role in role_categories}

    for e in gt_edges:
        for node in e:
            role = gt_roles.get(node)
            if role in gt_by_role:
                gt_by_role[role].add(normalize(node))

    for e in pred_edges:
        for node in e:
            role = pred_roles.get(node)
            if role in pred_by_role:
                pred_by_role[role].add(normalize(node))

    all_gt_nodes, all_pred_nodes = set(), set()
    total_tp = total_fp = total_fn = 0

    for role in role_categories:
        gt_nodes = list(gt_by_role[role])
        pred_nodes = list(pred_by_role[role])
        all_gt_nodes.update(gt_nodes)
        all_pred_nodes.update(pred_nodes)

        if not gt_nodes and not pred_nodes:
            continue

        if not gt_nodes or not pred_nodes:
            tp, fp, fn = 0, len(pred_nodes), len(gt_nodes)
        else:
            emb_gt = model.encode(gt_nodes, convert_to_tensor=True)
            emb_pred = model.encode(pred_nodes, convert_to_tensor=True)
            sim_matrix = util.pytorch_cos_sim(emb_pred, emb_gt).cpu().numpy()

            matched_gt = set()
            tp = 0
            for i, pred in enumerate(pred_nodes):
                j = sim_matrix[i].argmax()
                if sim_matrix[i][j] >= threshold and gt_nodes[j] not in matched_gt:
                    tp += 1
                    matched_gt.add(gt_nodes[j])
            fp = len(pred_nodes) - tp
            fn = len(gt_nodes) - tp

        precision = tp / (tp + fp) if tp + fp > 0 else 0.0
        recall = tp / (tp + fn) if tp + fn > 0 else 0.0
        f1 = 2 * precision * recall / (precision + recall) if precision + recall > 0 else 0.0
        jaccard = tp / (tp + fp + fn) if tp + fp + fn > 0 else 0.0

        role_metrics[role] = {
            "Precision": round(precision, 3),
            "Recall": round(recall, 3),
            "F1": round(f1, 3),
            "Jaccard": round(jaccard, 3),
            "TP": tp, "FP": fp, "FN": fn
        }

        total_tp += tp
        total_fp += fp
        total_fn += fn

    precision = total_tp / (total_tp + total_fp) if total_tp + total_fp > 0 else 0.0
    recall = total_tp / (total_tp + total_fn) if total_tp + total_fn > 0 else 0.0
    f1 = 2 * precision * recall / (precision + recall) if precision + recall > 0 else 0.0
    jaccard = total_tp / (total_tp + total_fp + total_fn) if total_tp + total_fp + total_fn > 0 else 0.0

    role_metrics["Overall"] = {
        "Precision": round(precision, 3),
        "Recall": round(recall, 3),
        "F1": round(f1, 3),
        "Jaccard": round(jaccard, 3),
        "TP": total_tp, "FP": total_fp, "FN": total_fn
    }

    return role_metrics

def export_node_metrics_to_excel(metrics_dict, method, prompt, domain, model, central_event, file_path=CSV_LOG):
    df = pd.DataFrame.from_dict(metrics_dict, orient="index").reset_index().rename(columns={"index": "Role"})
    df["Method"] = method
    df["Prompt"] = prompt
    df["Domain"] = domain
    df["Model"] = model
    df["Central Event"] = central_event

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if "Node_Metrics" in wb.sheetnames:
        ws = wb["Node_Metrics"]
        is_empty = ws.max_row <= 1
    else:
        ws = wb.create_sheet("Node_Metrics")
        is_empty = True

    if not is_empty:
        ws.append([])

    for r in dataframe_to_rows(df, index=False, header=is_empty):
        ws.append(r)

    wb.save(file_path)

    return df

def update_csv(row, file_path=CSV_LOG):
    import pandas as pd
    from openpyxl import Workbook
    import openpyxl

    ged_df = pd.DataFrame([row])
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "GED_Metrics" in wb.sheetnames:
        ws1 = wb["GED_Metrics"]
        is_empty = ws1.max_row <= 1
    else:
        ws1 = wb.create_sheet("GED_Metrics")
        is_empty = True

    if not is_empty:
        ws1.append([])

    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(ged_df, index=False, header=is_empty):
        ws1.append(r)

    # (Optional) Write Node_Mappings sheet here if needed

    wb.save(file_path)
    return pd.read_excel(file_path, sheet_name="GED_Metrics")



def get_cosine_node_mapping(gt_nodes, pred_nodes, threshold=0.8):
    gt_nodes = [normalize(n) for n in gt_nodes]
    pred_nodes = [normalize(n) for n in pred_nodes]

    emb_gt = model.encode(gt_nodes, convert_to_tensor=True)
    emb_pred = model.encode(pred_nodes, convert_to_tensor=True)

    sim_matrix = util.pytorch_cos_sim(emb_pred, emb_gt).cpu().numpy()

    pred_to_gt = {}
    matched_gt = set()
    mapping_scores = []

    for i, pred in enumerate(pred_nodes):
        j = sim_matrix[i].argmax()
        max_sim = sim_matrix[i][j]
        gt_label = gt_nodes[j]

        if max_sim >= threshold and gt_label not in matched_gt:
            pred_to_gt[pred] = gt_label
            matched_gt.add(gt_label)
            match_status = "Match"
        else:
            match_status = "No Match"

        mapping_scores.append({
            "Predicted": pred,
            "GT Match": gt_label if match_status == "Match" else "No Match",
            "Cosine Score": round(max_sim, 3),
            "Best GT Candidate": gt_label,
            "Best Score": round(max_sim, 3)
        })

    return pred_to_gt, mapping_scores


def compute_edge_metrics(gt_edges, pred_edges, gt_roles, pred_roles, threshold=0.8):
    metrics = {}
    overall_tp = overall_fp = overall_fn = 0

    gt_roles = {normalize(k): v for k, v in gt_roles.items()}
    pred_roles = {normalize(k): v for k, v in pred_roles.items()}

    gt_nodes = list(gt_roles.keys())
    pred_nodes = list(pred_roles.keys())

    node_mapping, _ = get_cosine_node_mapping(gt_nodes, pred_nodes, threshold)

    def edge_type(a, b):
        src_role = gt_roles.get(a) or pred_roles.get(a, "Unknown")
        tgt_role = gt_roles.get(b) or pred_roles.get(b, "Unknown")
        return f"{src_role}->{tgt_role}"

    norm_gt = [(normalize(a), normalize(b)) for a, b in gt_edges]
    norm_gt_set = set(norm_gt)

    remapped_pred_edges = []
    for a, b in pred_edges:
        a_norm, b_norm = normalize(a), normalize(b)
        remap_a = node_mapping.get(a_norm, a_norm)
        remap_b = node_mapping.get(b_norm, b_norm)
        remapped_pred_edges.append((remap_a, remap_b))

    gt_classified = {(a, b): edge_type(a, b) for a, b in norm_gt}
    pred_classified = {(a, b): edge_type(a, b) for a, b in remapped_pred_edges}

    for edge, etype in pred_classified.items():
        metrics.setdefault(etype, {"TP": 0, "FP": 0, "FN": 0})
        if edge in gt_classified and gt_classified[edge] == etype:
            metrics[etype]["TP"] += 1
            overall_tp += 1
        else:
            metrics[etype]["FP"] += 1
            overall_fp += 1

    for edge, etype in gt_classified.items():
        if edge not in pred_classified or pred_classified[edge] != etype:
            metrics.setdefault(etype, {"TP": 0, "FP": 0, "FN": 0})
            metrics[etype]["FN"] += 1
            overall_fn += 1

    rows = []
    for etype, vals in metrics.items():
        tp, fp, fn = vals["TP"], vals["FP"], vals["FN"]
        prec = tp / (tp + fp) if tp + fp > 0 else 0.0
        rec = tp / (tp + fn) if tp + fn > 0 else 0.0
        f1 = 2 * prec * rec / (prec + rec) if prec + rec > 0 else 0.0
        jacc = tp / (tp + fp + fn) if tp + fp + fn > 0 else 0.0
        rows.append({
            "Edge Type": etype,
            "Precision": round(prec, 3),
            "Recall": round(rec, 3),
            "F1": round(f1, 3),
            "Jaccard": round(jacc, 3),
            "TP": tp, "FP": fp, "FN": fn
        })

    # Overall row
    prec = overall_tp / (overall_tp + overall_fp) if overall_tp + overall_fp > 0 else 0.0
    rec = overall_tp / (overall_tp + overall_fn) if overall_tp + overall_fn > 0 else 0.0
    f1 = 2 * prec * rec / (prec + rec) if prec + rec > 0 else 0.0
    jacc = overall_tp / (overall_tp + overall_fp + overall_fn) if overall_tp + overall_fp + overall_fn > 0 else 0.0
    rows.append({
        "Edge Type": "Overall",
        "Precision": round(prec, 3),
        "Recall": round(rec, 3),
        "F1": round(f1, 3),
        "Jaccard": round(jacc, 3),
        "TP": overall_tp, "FP": overall_fp, "FN": overall_fn
    })

    return rows

def export_edge_metrics_to_excel(rows, method, prompt, domain, model, central_event, file_path=CSV_LOG):
    df = pd.DataFrame(rows)
    df["Method"] = method
    df["Prompt"] = prompt
    df["Domain"] = domain
    df["Model"] = model
    df["Central Event"] = central_event

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if "Edge_Metrics" in wb.sheetnames:
        ws = wb["Edge_Metrics"]
        is_empty = ws.max_row <= 1
    else:
        ws = wb.create_sheet("Edge_Metrics")
        is_empty = True

    if not is_empty:
        ws.append([])

    for r in dataframe_to_rows(df, index=False, header=is_empty):
        ws.append(r)

    wb.save(file_path)
    return df


def export_detailed_edge_comparisons(gt_edges, pred_edges, gt_roles, pred_roles, method, prompt, domain, model, central_event, file_path="ged_results.xlsx", threshold=0.8):
    gt_roles = {normalize(k): v for k, v in gt_roles.items()}
    pred_roles = {normalize(k): v for k, v in pred_roles.items()}

    gt_set = set((normalize(a), normalize(b)) for a, b in gt_edges)

    gt_nodes = list(gt_roles.keys())
    pred_nodes = list(pred_roles.keys())
    node_mapping, _ = get_cosine_node_mapping(gt_nodes, pred_nodes, threshold)

    # Remap predicted nodes
    remapped_pred_set = set()
    raw_pred_map = []
    for a, b in pred_edges:
        a_norm, b_norm = normalize(a), normalize(b)
        remap_a = node_mapping.get(a_norm, a_norm)
        remap_b = node_mapping.get(b_norm, b_norm)
        remapped_pred_set.add((remap_a, remap_b))
        raw_pred_map.append(((a_norm, b_norm), (remap_a, remap_b)))

    detailed_rows = []

    for a, b in gt_set:
        if (a, b) in remapped_pred_set:
            role_a = gt_roles.get(a) or pred_roles.get(a, "Unknown")
            role_b = gt_roles.get(b) or pred_roles.get(b, "Unknown")
            edge_type = f"{role_a}->{role_b}"
            detailed_rows.append({"Source": a, "Target": b, "Edge Type": edge_type, "Match Status": "TP"})
        else:
            role_a = gt_roles.get(a) or pred_roles.get(a, "Unknown")
            role_b = gt_roles.get(b) or pred_roles.get(b, "Unknown")
            edge_type = f"{role_a}->{role_b}"
            detailed_rows.append({"Source": a, "Target": b, "Edge Type": edge_type, "Match Status": "FN"})

    for (a_orig, b_orig), (a, b) in raw_pred_map:
        if (a, b) not in gt_set:
            role_a = pred_roles.get(a) or gt_roles.get(a, "Unknown")
            role_b = pred_roles.get(b) or gt_roles.get(b, "Unknown")
            edge_type = f"{role_a}->{role_b}"
            detailed_rows.append({"Source": a_orig, "Target": b_orig, "Edge Type": edge_type, "Match Status": "FP"})

    for row in detailed_rows:
        row["Method"] = method
        row["Prompt"] = prompt
        row["Domain"] = domain
        row["Model"] = model
        row["Central Event"] = central_event

    df = pd.DataFrame(detailed_rows)

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if "Edge_Comparisons" in wb.sheetnames:
        ws = wb["Edge_Comparisons"]
        is_empty = ws.max_row <= 1
    else:
        ws = wb.create_sheet("Edge_Comparisons")
        is_empty = True

    if not is_empty:
        ws.append([])

    for r in dataframe_to_rows(df, index=False, header=is_empty):
        ws.append(r)

    wb.save(file_path)
    return df


def export_node_comparisons_to_excel(gt_nodes, pred_nodes, gt_roles, pred_roles,
                                      method, prompt, model_name, domain,
                                      central_event, file_path="ged_results.xlsx", threshold=0.8):
    norm_gt_nodes = [normalize(n) for n in gt_nodes]
    norm_pred_nodes = [normalize(n) for n in pred_nodes]

    emb_gt = model.encode(norm_gt_nodes, convert_to_tensor=True)
    emb_pred = model.encode(norm_pred_nodes, convert_to_tensor=True)

    sim_matrix = util.pytorch_cos_sim(emb_pred, emb_gt).cpu().numpy()
    matched_gt = set()
    rows = []

    for i, pred_label in enumerate(norm_pred_nodes):
        j = sim_matrix[i].argmax()
        max_sim = sim_matrix[i][j]
        gt_label = norm_gt_nodes[j]

        if max_sim >= threshold and gt_label not in matched_gt:
            matched_gt.add(gt_label)
            gt_role = gt_roles.get(gt_label, "")
            pred_role = pred_roles.get(pred_label, "")
            match_status = "TP" if gt_role == pred_role else "Role Mismatch"
        else:
            gt_label = ""
            gt_role = ""
            pred_role = pred_roles.get(pred_label, "")
            match_status = "FP"

        rows.append({
            "Node Label": pred_label,
            "Ground Truth Role": gt_role,
            "Predicted Role": pred_role,
            "Match Status": match_status,
            "Method": method,
            "Prompt": prompt,
            "Model": model_name,
            "Domain": domain,
            "Central Event": central_event
        })

    for i, gt_label in enumerate(norm_gt_nodes):
        if gt_label not in matched_gt:
            rows.append({
                "Node Label": gt_label,
                "Ground Truth Role": gt_roles.get(gt_label, ""),
                "Predicted Role": "",
                "Match Status": "FN",
                "Method": method,
                "Prompt": prompt,
                "Model": model_name,
                "Domain": domain,
                "Central Event": central_event
            })

    df = pd.DataFrame(rows)

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if "Node_Comparisons" in wb.sheetnames:
        ws = wb["Node_Comparisons"]
        is_empty = ws.max_row <= 1
    else:
        ws = wb.create_sheet("Node_Comparisons")
        is_empty = True

    if not is_empty:
        ws.append([])

    for r in dataframe_to_rows(df, index=False, header=is_empty):
        ws.append(r)

    wb.save(file_path)
    return df


def export_cosine_mappings_to_ged_sheet(mapping_scores, method, prompt, domain, model, central_event, file_path="ged_results.xlsx"):
    df = pd.DataFrame(mapping_scores)
    df["Method"] = method
    df["Prompt"] = prompt
    df["Domain"] = domain
    df["Model"] = model
    df["Central Event"] = central_event

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    sheet_name = "Cosine_Remapping"

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_empty = ws.max_row <= 1
    else:
        ws = wb.create_sheet(sheet_name)
        is_empty = True

    if not is_empty:
        ws.append([])

    for r in dataframe_to_rows(df, index=False, header=is_empty):
        ws.append(r)

    wb.save(file_path)
    return df
